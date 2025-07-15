from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border, Font
from rest_framework import generics

from .models import (
    CourseRegistration,
    Course,
    University,
    StudyMode,
    ClassTimeOption,
)
from .serializers import (
    CourseRegistrationSerializer,
    CourseListSerializer,
    UniversitySerializer,
    StudyModeSerializer,
    ClassTimeOptionSerializer,
)


# --- Select APIs for dropdowns ---
class CourseListView(generics.ListAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseListSerializer
    permission_classes = []


class UniversityListView(generics.ListAPIView):
    queryset = University.objects.all()
    serializer_class = UniversitySerializer
    permission_classes = []


class StudyModeListView(generics.ListAPIView):
    queryset = StudyMode.objects.all()
    serializer_class = StudyModeSerializer
    permission_classes = []


class ClassTimeOptionListView(generics.ListAPIView):
    queryset = ClassTimeOption.objects.all()
    serializer_class = ClassTimeOptionSerializer
    permission_classes = []


# --- Registration Create API ---
class CourseRegistrationCreateView(generics.CreateAPIView):
    queryset = CourseRegistration.objects.all()
    serializer_class = CourseRegistrationSerializer
    permission_classes = []


# --- Registration List API + Excel Export ---
class CourseRegistrationListView(generics.ListAPIView):
    queryset = CourseRegistration.objects.all().prefetch_related(
        'course', 'intended_universities', 'ideal_class_times'
    ).select_related('preferred_study_mode').order_by('-created_at')
    serializer_class = CourseRegistrationSerializer
    permission_classes = []

    def list(self, request, *args, **kwargs):
        if request.query_params.get('export') == 'excel':
            return self.export_to_excel()
        return super().list(request, *args, **kwargs)

    def export_to_excel(self):
        registrations = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"

        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "Full Name", "Preferred Name", "Email", "Telegram", "City/Region",
            "DOB", "Education Level", "Courses", "Intended Universities",
            "Study Mode", "Ideal Times", "Top Goal", "Merit Discount?",
            "Heard About Us", "Wants Premium", "Notes", "Created At"
        ]
        ws.append(headers)
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 25

        for row_num, reg in enumerate(registrations, start=2):
            row = [
                reg.full_name,
                reg.preferred_name,
                reg.email,
                reg.telegram_handle,
                reg.current_city_or_region,
                reg.date_of_birth.strftime('%Y-%m-%d') if reg.date_of_birth else "",
                reg.current_education_level,
                ", ".join(c.name for c in reg.course.all()),
                ", ".join(u.name for u in reg.intended_universities.all()),
                reg.preferred_study_mode.name if reg.preferred_study_mode else "",
                ", ".join(t.label for t in reg.ideal_class_times.all()),
                reg.top_goal,
                "Yes" if reg.apply_for_merit_discount else "No",
                reg.heard_about_us,
                "Yes" if reg.wants_premium else "No",
                reg.notes or "",
                reg.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical='top')
                cell.border = thin_border

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=course_registrations.xlsx'
        wb.save(response)
        return response
